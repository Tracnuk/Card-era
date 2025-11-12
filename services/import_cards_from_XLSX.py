import sys
import os

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import load_workbook

from repositories.card_db_repository import CardDbRepository
from models.card import Card

card_db_storage = CardDbRepository()

class ImportXLSXService:
    def import_data(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        xlsx_path = os.path.join(script_dir, 'cards.xlsx')        
        try:
            workbook = load_workbook(xlsx_path)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            for row_idx, row in enumerate(rows, start=2):
                if len(row) < 9:
                    print(f"Пропущена неполная строка (строка {row_idx}): {row}")
                    continue
                card_object = Card(
                    rarity=str(row[0]),
                    type_card=str(row[1]),
                    name=str(row[2]),
                    count=int(row[3]) if row[3] is not None else 0,
                    hp=int(row[4]),
                    damage=int(row[5]),
                    energy=int(row[6]),
                    price=float(row[7]) if row[7] is not None else 0.0,
                    link_of_picture=str(row[8]) if row[8] is not None else ""
                )
                card_db_storage.add_card(card_object)
            print("Импорт данных успешно завершён.")
        except FileNotFoundError:
            print(f"Ошибка: файл не найден по пути: {xlsx_path}")
        except PermissionError:
            print(f"Ошибка: нет доступа к файлу (возможно, он открыт в Excel): {xlsx_path}")
        except ValueError as e:
            print(f"Ошибка преобразования данных в строке {row_idx}: {e}")
        except Exception as e:
            print(f"Неожиданная ошибка при обработке файла: {e}")
